from executePing import execPing
from getExcelColumnHeaders import getColumnHeaders
import openpyxl


def pingColumn(ws1, columnHeader, files1, reset):
    if reset:
        permissions = "w"
    else:
        permissions = "a"

    successes = open(files1["successFile"], permissions)
    timeouts = open(files1["timeoutFile"], permissions)
    failures = open(files1["failuresFile"], permissions)

    headers = getColumnHeaders(ws1)  # header dictionary

    for i in range(2, ws1.max_row + 1):
        cell = (ws1.cell(row=i, column=headers[columnHeader]))

        sep = "=========================================================" + "\n"
        if cell.value != columnHeader and cell.value is not None:
            ping = execPing(hostName=cell.value, attempts="2", waitTime="0.2")
            if ping[0] == "Success: ":
                successes.write(ping[1] + "\n")
                successes.write(sep)
                print(sep)
                print(ping[1] + "\n")
            elif ping[0] == "Request timeout: ":
                timeouts.write(ping[1] + "\n")
                timeouts.write(sep)
                print(sep)
                print(ping[1] + "\n")
            elif ping[0] == "Failed: ":
                failures.write("*" + cell.value + "*" + ": " + ping[1] + "\n")
                failures.write(sep)
                print(sep)
                print("*" + cell.value + "*" + ": " + ping[1] + "\n")


if __name__ == "__main__":
    inFileName = "exampleWorkBook.xlsx"
    # inFileName = "новые 5 очередь РСМОБ.xlsm"
    wb = openpyxl.load_workbook(inFileName, read_only=True)  # load workbook
    ws = wb.get_sheet_by_name("Sheet1")  # select worksheet
    # ws = wb["Камеры"]  # select worksheet
    # ws = wb["ping"]  # select worksheet

    cp = "_pings_from_"
    files = {"successFile": "successful" + cp + inFileName.replace(".xlsx", ".txt"),
             "timeoutFile": "timeout" + cp + inFileName.replace(".xlsx", ".txt"),
             "failuresFile": "failure" + cp + inFileName.replace(".xlsx", ".txt")}

    # pingColumn(ws1=ws, columnHeader="HostName", files1=files, reset=True)
    # pingColumn(ws1=ws, columnHeader="IP", files1=files, reset=True)
    pingColumn(ws1=ws, columnHeader="IP", files1=files, reset=True)
