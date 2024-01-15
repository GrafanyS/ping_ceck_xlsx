import openpyxl


def getColumnHeaders(ws):
    headers = {}  # dict to store headers
    try:  # make sure worksheet is valid
        for i in range(1, ws.max_column + 1):
            headers[str(ws.cell(row=1, column=i).value)] = i

    except ValueError:
        print("Произошла ошибка при попытке" +
              " получить заголовки из файла excel.")
    return headers


def exampleUsage():
    wb = openpyxl.load_workbook("exampleWorkBook.xlsx", read_only=True)  # load workbook
    # wb = openpyxl.load_workbook("новые 5 очередь РСМОБ.xlsm", read_only=True)  # load workbook
    ws = wb.get_sheet_by_name("Sheet1")  # select worksheet
    # ws = wb["ping"]  # select worksheet
    # ws = wb["Камеры"]  # select worksheet
    headers = getColumnHeaders(ws)  # get list of headers

    try:  # check if requested header actually exists
        cell = (ws.cell(row=2, column=headers['IP']))
        print(str(cell.value))

    except ValueError:
        print("Этот заголовок, по-видимому, не существует.")


if __name__ == "__main__":
    exampleUsage()
